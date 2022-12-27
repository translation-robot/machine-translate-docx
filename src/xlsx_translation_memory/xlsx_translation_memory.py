from openpyxl import load_workbook
import datetime
import timeit
import re
from typing import Any
from typing import NamedTuple
from newmm_tokenizer.tokenizer import word_tokenize
import tinysegmenter
from pprint import pprint
import traceback
import os

# - *- coding: utf- 8 - *-

class SearchAndReplace(NamedTuple):
    """This class is used as a structure to contain a search and replace enty as a translation memory
    it is used in a list of translation memory contained in an excel file for object xlsx_translation_memory"""
    Search: str
    SearchRegularExpression: str
    Replace: str
    RegularExpression: bool
    CaseSensitive: bool
    IgnoreWordBoundary: bool
    KeepSpacesAtTheEndOfSearch: bool
    RegularExpressionCompiled: Any
    FollowedByAnotherWord : bool
    number_replacement: int
    number_do_not_split: int

class xlsx_translation_memory():
    """class xlsx_translation_memory
    is a class that read excel xlsx files and build a list of search and replace
    with the following columns:
    Search	Replace	RegularExpression	CaseSensitive	IgnoreWordBoundary	KeepSpacesAtTheEndOfSearch
    Column Search is mandatory, column replace is optional
    columns CaseSensitive, IgnoreWordBoundary and KeepSpacesAtTheEndOfSearch are optional
    and defaults to False (F), to assign to True, use T value un excel cell."""
    def string_is_yes(self, string):
        """This method return True if a string's first non blank character is letter Y (case insensitive)"""
        res = False
        first_char = ''
        if string is not None:
            string_1st_char = string.strip()
            if len(string_1st_char) > 0:
                first_char = string_1st_char[0:1].upper()
        if first_char == 'Y':
            res = True
        return res

    def read_xlsx_search_and_replace(self):
        """"
        This method reads the content of the xlsx file
        and is called by method load_xlsx_translation_memory
        create a search and replace dictionary list for all the sheets"""
        try:
            print("Reading the content of the translation memory file %s" % (self.xlsx_path))
            
            self.ws = self.wb.active
            found_after_ws = False
            sheet_name = None
            
            search_multiple_blanks = r'[ \t]+'
            replace_single_blank = r' '
            re_multiple_blanks = re.compile(search_multiple_blanks)
            
            search_start_word_boundary = r'^\b.+'
            re_start_word_boundary = re.compile(search_start_word_boundary)
            
            search_end_word_boundary = r'.+\b$'
            re_end_word_boundary = re.compile(search_end_word_boundary)

            sheet_name = 'keep_on_same_line'
            if not bool(self.worksheets_search_and_replace_dictionary.get(sheet_name)):
                print("Sheet '%s' not found in excel file '%s'" % (sheet_name, self.xlsx_path))

            for s in range(len(self.wb.sheetnames)):
                self.wb.active = s
                self.ws = self.wb.active
                sheet_name = self.wb.sheetnames[s].lower().strip()
                	
                # We create an empty list of search and replace
                # for the sheet in the dictionary
                self.worksheets_search_and_replace_dictionary[sheet_name] = []
                search_and_replace_list = []

                rowno = 0
                for row in self.ws.values:
                    #print ("len=%s" % (len(row)))
                    rowno = rowno + 1
                    if rowno == 1:
                        continue
                    colpos = 0
                    while colpos < len(row):
                        #if (row[colpos] is not None):
                        #    #print("cell[%i,%i]=%s" % (rowno, colpos, row[colpos]))
                        colpos += 1
                    
                    Search = row[0]
                    if Search is None:
                        Search = ''
                        continue
                    else:
                        subn_result = re_multiple_blanks.subn(replace_single_blank, Search)
                        subn_count = subn_result[1]
                        Search = subn_result[0]
                    
                    Replace = row[1]
                    if Replace is None:
                        Replace = ''
                    
                    KeepSpacesAtTheEndOfSearch = row[5]
                    if self.string_is_yes(KeepSpacesAtTheEndOfSearch):
                        KeepSpacesAtTheEndOfSearch = True
                    else:
                        KeepSpacesAtTheEndOfSearch = False
                        Search = Search.strip()
                        Replace = Replace.strip()

                    RegularExpression = row[2]
                    RegularExpressionCompiled = None
                    if self.string_is_yes(RegularExpression):
                        RegularExpression = True
                        try:
                            RegularExpressionCompiled = re.compile(Search)
                            SearchRegularExpression = Search
                        except:
                            RegularExpression = False
                            RegularExpressionCompiled = None
                            SearchRegularExpression = re.escape(Search)
                            var = traceback.format_exc()
                            print(var)
                    else:
                        RegularExpression = False
                        SearchRegularExpression = re.escape(Search)
                    
                    IgnoreWordBoundary = row[4]
                    if self.string_is_yes(IgnoreWordBoundary):
                        IgnoreWordBoundary = True
                    else:
                        IgnoreWordBoundary = False
                        
                        if re_start_word_boundary.match (Search) is not None:
                            SearchRegularExpression = "\\b%s" % (SearchRegularExpression)
                        
                        if re_end_word_boundary.match (Search) is not None:
                            SearchRegularExpression = "%s\\b" % (SearchRegularExpression)

                    try:
                        FollowedByAnotherWord = row[6]
                        if self.string_is_yes(FollowedByAnotherWord):
                            FollowedByAnotherWord = True
                            SearchRegularExpression = "%s[ \\t]+[^ \\t]+" % SearchRegularExpression
                        else:
                            FollowedByAnotherWord = False
                    except:
                        FollowedByAnotherWord = False

                    number_replacement = 0
                    number_do_not_split = 0
                    
                    CaseSensitive = row[3]
                    if self.string_is_yes(CaseSensitive):
                        CaseSensitive = True
                    else:
                        CaseSensitive = False
                        SearchRegularExpression = "(?i)%s" % (SearchRegularExpression)
                    
                    if Search is not None and rowno > 1:
                        try:
                            RegularExpressionCompiled = re.compile(SearchRegularExpression)
                        except:
                            RegularExpression = False
                            RegularExpressionCompiled = None
                            var = traceback.format_exc()
                            print(var)
                        
                        sr = SearchAndReplace(Search, SearchRegularExpression, Replace, RegularExpression, CaseSensitive, IgnoreWordBoundary,
                                              KeepSpacesAtTheEndOfSearch,RegularExpressionCompiled, FollowedByAnotherWord, number_replacement, number_do_not_split)
                        #if sheet_name == "keep_on_same_line":
                        #    print("search=%s" % (sr.Search))
                        search_and_replace_list.append(sr)
                    
                    self.worksheets_search_and_replace_dictionary[sheet_name] = search_and_replace_list
                    
        except Exception:
            print ("Error reading excel translation memory file.")
            var = traceback.format_exc()
            print("ERROR: %s" % (var))
            self.worksheets_search_and_replace_dictionary[sheet_name] = search_and_replace_list
    

    def search_and_replace_text(self, sheet_name, text_to_replace):
        """This method loop for all the search and replace list
        and does the search each in text_to_replace string"""
        nb_replacements = 0
        sheet_name = sheet_name.lower()

        try:
            if self.worksheets_search_and_replace_dictionary[sheet_name] is None:
                print("Excel worksheet %s was not found %s." % (sheet_name))
        except Exception:
            return text_to_replace, 0

        text_replaced = text_to_replace
        try:
            #print("Search and replace text '%s'" % (text_replaced))
            se_no = 1
            for search_replace_item in self.worksheets_search_and_replace_dictionary[sheet_name]:
                #print("Search '%s' and replace '%s' # %d" % (search_replace_item.Search , search_replace_item.Replace ,se_no)) 
                #print("Search and replace # %d" % (se_no)) 
                subn_result = search_replace_item.RegularExpressionCompiled.subn(search_replace_item.Replace, text_replaced)
                subn_count = subn_result[1]
                if subn_count > 0:
                    search_replace_item._replace (number_replacement = search_replace_item.number_replacement + subn_count)
                    self.total_number_of_replacements = self.total_number_of_replacements + subn_count
                    nb_replacements = nb_replacements + subn_count
                    #print ("Original text :%s" % (text_replaced)) 
                    print ("Replaced '%s' by '%s' %d times (%s)." % (search_replace_item.Search, search_replace_item.Replace, subn_count, sheet_name))
                    #print ("Replaced text :%s" % (subn_result[0])) 
                text_replaced = subn_result[0]
                
                search_replace_item = search_replace_item._replace (number_replacement = search_replace_item.number_replacement + subn_count)
                self.worksheets_search_and_replace_dictionary[sheet_name][se_no-1] = search_replace_item
                #print ("res_str=%s, count=%i" % (res_str, subn_count))
                #pprint (search_replace_item)
                se_no = se_no + 1
            if nb_replacements > 0:
                print ("\n")
        except Exception:
            print ("Error reading excel translation memory file.")
            var = traceback.format_exc()
            print("ERROR: %s" % (var))
        return text_replaced, nb_replacements

    def tokenize_phrase(self, text, dest_lang):
        len_text = 0
        try:
            len_text = len(text)
        except:
            print("Error getting text length.")
        do_not_split_array = [0] * (len_text)

        if dest_lang.lower() == 'ja' or dest_lang.lower() == 'zh-cn' or dest_lang.lower() == 'zh' or dest_lang.lower() == 'zh-tw' or dest_lang.lower() == 'ko':
            words = self.cjk_segmenter.tokenize(text)
        # In other languages, just use spaces
        elif dest_lang.lower() == 'th':
            # words = thai_segmenter(text)
            words = word_tokenize(text)
        # In other languages, just use spaces
        else:
            # Old simple tokenizer used split
            #words = text.split()

            #input("Setting do not split array values")
            try:
                # print("Search and replace text '%s'" % (text_replaced))
                se_no = 1
                sheet_name = "keep_on_same_line"
                nb_replacements = 0
                sheet_name = sheet_name.lower()
                text_replaced = text
                text_len = len(text)
                text_do_not_split_pos_array = [0] * (text_len)

                if not bool(self.worksheets_search_and_replace_dictionary.get(sheet_name)):
                    words = text.split()
                    return words

                se_no = 1
                for search_do_not_split_item in self.worksheets_search_and_replace_dictionary[sheet_name]:
                    # print("Search '%s' and replace '%s' # %d" % (search_do_not_split_item.Search , search_do_not_split_item.Replace ,se_no))
                    # print("Search and replace # %d" % (se_no))

                    iterator_do_not_replace_match = search_do_not_split_item.RegularExpressionCompiled.finditer(text)
                    nb_match_do_not_split = 0
                    #print("nb_match_do_not_split: %d" % (nb_match_do_not_split))
                    for match in iterator_do_not_replace_match:
                        nb_match_do_not_split = nb_match_do_not_split + 1

                        search_do_not_split_item = search_do_not_split_item._replace(
                            number_do_not_split=search_do_not_split_item.number_do_not_split + 1)
                        self.worksheets_search_and_replace_dictionary[sheet_name][se_no-1] = search_do_not_split_item
                        match_start_pos = match.start()
                        match_end_pos = match.end()
                        #print(match.span())
                        #print(match.start())
                        #print(match.end())
                        for char_pos_dont_split in range(match_start_pos, match_end_pos):
                            text_do_not_split_pos_array[char_pos_dont_split] = 1
                        #print("search : %s" % (search_do_not_split_item.Search))
                        #print("'" + text[match.start():match.end()] + "'")
                        #print(text_do_not_split_pos_array)
                        #input("Here in search do not split")
                    se_no = se_no + 1

                spaces_pattern = re.compile(r' +')
                iterator_spaces = spaces_pattern.finditer(text)
                split_text_array = []
                text_current_pos = 0
                for space_match in iterator_spaces:
                    spaces_match_start_pos = space_match.start()
                    spaces_match_end_pos = space_match.end()
                    #print(space_match.span())
                    #print(space_match.start())
                    #print(space_match.end())
                    #print("text_current_pos: %d" % (text_current_pos))
                    #print("text current: '%s'" % (text[text_current_pos:spaces_match_start_pos]))
                    #print(text_do_not_split_pos_array)
                    if text_do_not_split_pos_array[spaces_match_start_pos] == 0:
                        do_not_split_token = text[text_current_pos:spaces_match_start_pos]
                        #print("do_not_split_token : '%s'" % (do_not_split_token))
                        split_text_array.append(do_not_split_token)
                        text_current_pos = spaces_match_end_pos
                        #print("Remaining: '%s'" % (text[text_current_pos:]))
                        pass

                    #print("split_text_array:")
                    #print(split_text_array)
                if text_len > text_current_pos:
                    do_not_split_token = text[text_current_pos:]
                    #print("do_not_split_token : '%s'" % (do_not_split_token))
                    split_text_array.append(do_not_split_token)

                #print("text: %s" % (text))
                #print("split_text_array:")
                #print(split_text_array)
                #print(text_do_not_split_pos_array)
                words = split_text_array
            except Exception:
                print("Error reading excel translation memory file.")
                var = traceback.format_exc()
                print("ERROR: %s" % (var))
            #input("This function is not complete, use only for debug")
        return words

    def pprint_translation_memory_list(self):
        """Method pprint_translation_memory_list pprint the worksheets_search_and_replace_dictionary"""
        pprint (self.worksheets_search_and_replace_dictionary['before'])
        pprint (self.worksheets_search_and_replace_dictionary['after'])
        try:
            pprint (self.worksheets_search_and_replace_dictionary['keep_on_same_line'])
        except:
            pass

    def get_sheet_number_of_keep_on_same_line_matches(self, sheet_name):
        """This method prints the number of replaced search items
        and the number of it they were replaced"""
        nb_do_not_split = 0
        sheet_name = sheet_name.lower()
        try:
            #print("Search and replace text '%s'" % (text_replaced))
            se_no = 1
            for keep_on_same_line_item in self.worksheets_search_and_replace_dictionary[sheet_name]:
                nb_do_not_split= self.worksheets_search_and_replace_dictionary[sheet_name][se_no-1].number_do_not_split + nb_do_not_split
                se_no = se_no + 1
        except Exception:
            print ("Error in get_sheet_number_of_keep_on_same_line_matches.")
            var = traceback.format_exc()
            print("ERROR: %s" % (var))
        return nb_do_not_split

    def get_sheet_number_lines(self, sheet_name):
        """This method prints the number of lines in
        the excel sheet without counting the header line"""
        nb_lines = 0
        sheet_name = sheet_name.lower()

        try:
            nb_lines = len(self.worksheets_search_and_replace_dictionary[sheet_name])
        except Exception:
            # Ignore the missing excel sheet and leave the variable nb_lines to 0
            #print ("Warning, sheet %s does not exist in the excel file." % (sheet_name))
            var = traceback.format_exc()
            #print("Warning : %s" % (var))

        return nb_lines

    def get_sheet_number_of_replacements(self, sheet_name):
        """This method prints the number of replaced search items
        and the number of it they were replaced"""
        nb_replacements = 0
        sheet_name = sheet_name.lower()

        try:
            if self.worksheets_search_and_replace_dictionary[sheet_name] is None:
                #print("Excel worksheet %s was not found %s." % (sheet_name))
                pass
        except Exception:
            return 0

        try:
            #print("Search and replace text '%s'" % (text_replaced))
            se_no = 1
            for search_replace_item in self.worksheets_search_and_replace_dictionary[sheet_name]:
                nb_replacements = self.worksheets_search_and_replace_dictionary[sheet_name][se_no-1].number_replacement + nb_replacements
                se_no = se_no + 1
        except Exception:
            print ("Error in get_sheet_number_of_replacements.")
            var = traceback.format_exc()
            print("ERROR: %s" % (var))
        return nb_replacements


    def get_sheet_number_of_do_not_split_match(self, sheet_name):
        """This method prints the number of replaced search items
        and the number of it they were replaced"""
        nb_do_not_split_match = 0
        sheet_name = sheet_name.lower()

        try:
            if self.worksheets_search_and_replace_dictionary[sheet_name] is None:
                #print("Excel worksheet %s was not found %s." % (sheet_name))
                pass
        except Exception:
            return 0

        try:
            #print("Search and replace text '%s'" % (text_replaced))
            se_no = 1
            for do_not_spit_item in self.worksheets_search_and_replace_dictionary[sheet_name]:
                nb_do_not_split_match = self.worksheets_search_and_replace_dictionary[sheet_name][se_no-1].number_do_not_split + nb_do_not_split_match
                se_no = se_no + 1
        except Exception:
            print ("\nWarning in get_sheet_number_of_do_not_split_match, sheet '%s' not found in excel file '%s'."
                   % (sheet_name, self.xlsx_path))
            var = traceback.format_exc()
            print("This will be ignored." % (sheet_name))
        return nb_do_not_split_match

    def print_replaced_items_number_of_replacements(self, sheet_name):
        """This method Method print_replaced_items_number_of_replacements prints the number of replaced search items
        and the number of it they were replaced"""
        sheet_name = sheet_name.lower()

        try:
            if self.worksheets_search_and_replace_dictionary[sheet_name] is None:
                #print("Excel worksheet %s was not found %s." % (sheet_name))
                pass
        except Exception:
            return

        try:
            if self.wb is None or self.worksheets_search_and_replace_dictionary[sheet_name] is None:
                print ("\nTranslation search and replace file '%s' missing : ignored." % (self.xlsx_path))
            else:
                print ("\nSearch and replace number of replacements for excel sheet name '%s' :\n" % (sheet_name))
                se_no = 1
                for search_replace_item in self.worksheets_search_and_replace_dictionary[sheet_name]:
                    if search_replace_item.number_replacement > 0:
                        if search_replace_item.number_replacement > 1:
                            str_print =  "Replaced '%s' (excel line %d), by '%s' %d times."
                        else:
                            str_print =  "Replaced '%s' (excel line %d), by '%s' %d time."
                        print (str_print % (search_replace_item.Search , se_no + 1, search_replace_item.Replace, search_replace_item.number_replacement))
                    se_no = se_no + 1
                number_of_replacements = self.get_sheet_number_of_replacements(sheet_name)
                print ("\nNumber of replacements using sheet '%s' : %d from a list of %d entries.\n" % (sheet_name, number_of_replacements, len(self.worksheets_search_and_replace_dictionary[sheet_name])))
        except Exception:
            print ("Error calling methof print_replaced_items_number_of_replacements")
            var = traceback.format_exc()
            print("ERROR: %s" % (var))

    def print_do_not_split_number_of_matches(self, sheet_name):
        """This method Method print_replaced_items_number_of_replacements prints the number of replaced search items
        and the number of it they were replaced"""
        sheet_name = sheet_name.lower()
        total_number_do_not_split = 0
        try:
            if self.wb is None or self.worksheets_search_and_replace_dictionary[sheet_name] is None:
                print ("\nTranslation search and replace file '%s' missing : ignored." % (self.xlsx_path))
            else:
                print ("\nCharacter string do not split number matches '%s' :\n" % (sheet_name))
                se_no = 1

                for do_not_split_item in self.worksheets_search_and_replace_dictionary[sheet_name]:
                    if do_not_split_item.number_do_not_split > 0:
                        total_number_do_not_split = total_number_do_not_split + do_not_split_item.number_do_not_split
                        if do_not_split_item.FollowedByAnotherWord == True:
                            followed_by_another_word = ", followed by another word, "
                        else:
                            followed_by_another_word = ""
                        if do_not_split_item.number_do_not_split > 1:
                            str_print =  "Keep on same line '%s'%s(excel line %d) matched %d times."
                        else:
                            str_print =  "Keep on same line '%s' %s(excel line %d) matched %d time."
                        print (str_print % (do_not_split_item.Search , followed_by_another_word, se_no + 1, do_not_split_item.number_do_not_split))
                    se_no = se_no + 1

                number_do_not_split = self.get_sheet_number_of_keep_on_same_line_matches(sheet_name)

                print ("\nNumber of keep on the same line using sheet '%s' : %d from a list of %d entries.\n" % (
                    sheet_name, total_number_do_not_split, len(self.worksheets_search_and_replace_dictionary[sheet_name])))
        except Exception:
            #print ("Warning calling method print_replaced_items_number_of_replacements, sheet %s was not found." % (sheet_name))
            var = traceback.format_exc()
            #print("ERROR: %s" % (var))

    def load_xlsx_translation_memory(self, xlsx_path):
        """This method opens the xlsx file"""
        try:
            self.xlsx_path = xlsx_path
            self.worksheets_search_and_replace_dictionary = {}
            self.wb = load_workbook(self.xlsx_path)
            self.ws = self.wb.active
            self.read_xlsx_search_and_replace()
        except Exception:
            print ("Error loading excel translation memory file.")
            var = traceback.format_exc()
            print("ERROR: %s" % (var))
            self.wb = None
            self.ws = None

    def __init__(self, xlsx_path):
        """""""The constructor of class xlsx_translation_memory"""
        self.xlsx_path = xlsx_path
        self.total_number_of_replacements = 0
        self.worksheets_search_and_replace_dictionary = {}

        self.cjk_segmenter = tinysegmenter.TinySegmenter()

        print ("Init XLSXTranslationMemory")
        if xlsx_path is None :
            self.wb = None
            self.ws = None
            self.worksheets_search_and_replace_dictionary = {}
        else:
            try:
                if not os.path.exists(xlsx_path):
                    print ("ERROR: File not found: %s" % (xlsx_path))
                else:
                    self.xlsx_splitted_filename = os.path.splitext(os.path.basename(xlsx_path))

                    # number of segment separated by dot in the docx filename
                    self.xlsx_splitted_filename_size = len(self.xlsx_splitted_filename)

                    if self.xlsx_splitted_filename_size > 1:
                        self.xlsx_file_to_translate_extension = self.xlsx_splitted_filename[
                            self.xlsx_splitted_filename_size - 1].lower()

                    if self.xlsx_file_to_translate_extension == ".xlsx":
                        self.load_xlsx_translation_memory(xlsx_path)
                    else:
                        print("ERROR: translation memory file is not xlsx file: %s" % (xlsx_path))
                        self.wb = None
                        self.ws = None
            except Exception:
                self.wb = None
                self.ws = None
                self.worksheets_search_and_replace_dictionary = {}

def main():
    start = timeit.timeit()

    xtm = xlsx_translation_memory('wordlist_fr_v1.xlsx')
    res1 = xtm.search_and_replace_text("Bonjour Principal, quelle belle journée, Principal ! C'est fantastique !")
    res2 = xtm.search_and_replace_text("Fantastique !")
    print ("res1 = %s" % (res1))
    print ("res2 = %s" % (res2))
    #xtm.pprint_translation_memory_list()
    print ("total_number_of_replacements = %d" % (xtm.total_number_of_replacements))
	
    end = timeit.timeit()
    elapsedtime = end - start
    print ("\nXlsx file read. Elasped time: %s (h:mm:ss.mmm)" % str(datetime.timedelta(seconds=(end - start))))

if __name__ == '__main__':
    main()
